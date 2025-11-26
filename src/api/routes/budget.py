from typing import Optional
from fastapi import APIRouter, File, Query, UploadFile, status, HTTPException, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from src.models.branch_list import BranchListOut
import pandas as pd
import calendar
from datetime import timedelta
from io import BytesIO
from src.models.budget import defaultBudgetModel
from src.models.projection import ProjectionEstimateIn, ProjectionInputModel
from src.models.projected_allocation_branch import BranchTotalsIn, BranchTotalsOut
from src.models.projected_allocation_branch_monthly import BranchMonthlyTotalsIn, BranchMonthlyTotalsOut
from src.services.budget import calculateDefault
from src.services.smart_ramadan import SmartRamadanSystem
from src.services.saveProjections import upsert_projection_estimate, upsert_projection_input, upsert_projection_input_batch, upsert_projection_estimate_batch
from src.services.budget_state import compute_or_reuse
import traceback
from src.models.projection_est_adjust import AllocateGrandTotalIn, AllocateGrandTotalOut
from src.services.projected_allocation import allocate_grand_total
from src.models.projected_allocation_monthly import MonthlyTotalsIn, MonthlyTotalsOut
from src.services.allocation_service_monthly import allocate_monthly_totals
from src.services.allocation_service_branch import allocate_branch_totals
from src.services.allocation_service_branch_monthly import allocate_branch_monthly_totals
from src.services.branch_service import list_branches
from src.services.importdata import validate_sales_csv

budgetRouter = APIRouter(prefix="/api")

# Import authentication dependency
from src.api.routes.auth import get_current_user


def filter_budget_data_by_permissions(data: dict, user: dict) -> dict:
    """Filter budget data based on user's brand and branch permissions"""
    # Check if user is super_admin (has access to all data)
    is_super_admin = any(role.get("name") == "super_admin" for role in user.get("roles", []))
    
    if is_super_admin:
        # Super admin sees all data
        return data
    
    # Get user's brand and branch access
    user_brand_ids = user.get("brand_access", [])
    user_branch_ids = user.get("branch_access", [])
    
    # If user has no restrictions, return all data
    if not user_brand_ids and not user_branch_ids:
        return data
    
    # Filter the data
    filtered_data = data.copy()
    
    # Filter brands if data has a 'data' key with brands
    if "data" in filtered_data and isinstance(filtered_data["data"], list):
        filtered_brands = []
        
        for brand in filtered_data["data"]:
            brand_id = brand.get("brand_id")
            
            # Check if user has access to this brand
            if user_brand_ids and brand_id not in user_brand_ids:
                continue  # Skip this brand
            
            # Filter branches within the brand
            if user_branch_ids and "branches" in brand:
                filtered_branches_list = [
                    branch for branch in brand["branches"]
                    if branch.get("branch_id") in user_branch_ids
                ]
                
                # Only include brand if it has accessible branches
                if filtered_branches_list:
                    filtered_brand = brand.copy()
                    filtered_brand["branches"] = filtered_branches_list
                    filtered_brands.append(filtered_brand)
            else:
                # No branch filtering needed, include whole brand
                filtered_brands.append(brand)
        
        filtered_data["data"] = filtered_brands
    
    return filtered_data




@budgetRouter.post("/calculatedefault", status_code=status.HTTP_202_ACCEPTED)
def defaultCalculation(request: defaultBudgetModel, current_user: dict = Depends(get_current_user)):
    try:
        # res=calculateDefault(request)
        # return {"Data":res}
        result = compute_or_reuse(request, recompute=calculateDefault)
        # Filter results based on user permissions
        return filter_budget_data_by_permissions(result, current_user)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@budgetRouter.get("/brands-list", status_code=status.HTTP_200_OK)
def get_brands_list(current_user: dict = Depends(get_current_user)):
    """Lightweight endpoint to get just brands and branches list without calculations"""
    try:
        from src.core.db import get_session, close_session
        from src.db.dbtables import Brand, Branch
        
        dbs = get_session()
        try:
            # Query brands and branches
            brands_query = dbs.query(Brand).filter(Brand.is_deleted == False).order_by(Brand.id).all()
            
            # Build simple structure
            brands_data = []
            for brand in brands_query:
                branches = dbs.query(Branch).filter(
                    Branch.brand_id == brand.id,
                    Branch.is_deleted == False
                ).order_by(Branch.id).all()
                
                brand_dict = {
                    "brand_id": brand.id,
                    "brand_name": brand.name,
                    "branches": [
                        {
                            "branch_id": branch.id,
                            "branch_name": branch.name
                        }
                        for branch in branches
                    ]
                }
                brands_data.append(brand_dict)
            
            result = {
                "data": brands_data,
                "config": {
                    "compare_year": 2025,  # CY - Base year (actual data for averages)
                    "budget_year": 2026    # BY - Budget year (future estimates)
                }
            }
            
            # Filter by permissions
            return filter_budget_data_by_permissions(result, current_user)
            
        finally:
            close_session(dbs)
            
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch brands list: {str(e)}"
        )


@budgetRouter.post("/weekend-effect", status_code=status.HTTP_200_OK)
def get_weekend_effect(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Calculate weekend effect for selected branches (matches home page calculation)
    Request body: {
        "branch_ids": [1, 2, 3],
        "compare_year": 2023,
        "budget_year": 2025
    }
    """
    try:
        import pandas as pd
        import calendar
        from datetime import datetime
        from src.core.db import get_session, close_session
        from src.db.dbtables import Branch
        
        branch_ids = request.get("branch_ids", [])
        compare_year = request.get("compare_year", 2025)  # Default to 2025 (CY - base year)
        budget_year = request.get("budget_year", 2026)   # Default to 2026 (BY - budget year)
        
        if not branch_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="branch_ids is required"
            )
        
        # Load BaseData.pkl
        BASE_DATA_PATH = "/home/user/backend/Backend/BaseData.pkl"
        df = pd.read_pickle(BASE_DATA_PATH)
        
        # Ensure business_date is datetime
        df['business_date'] = pd.to_datetime(df['business_date'])
        
        # Filter by branch_ids
        df = df[df['branch_id'].isin(branch_ids)]
        
        # Filter by years (compare year only for actual data)
        df_compare = df[df['business_date'].dt.year == compare_year].copy()
        
        # Add weekday info
        df_compare['month'] = df_compare['business_date'].dt.month
        df_compare['day_name'] = df_compare['business_date'].dt.day_name()
        
        # Helper function to count calendar occurrences (matches home page logic)
        def count_day_occurrences(year, month, day_number):
            """Count how many times a weekday occurs in a month using calendar"""
            month_days = calendar.monthcalendar(year, month)
            return sum(1 for week in month_days if week[day_number] != 0)
        
        # Generate day occurrence lookup (matches home page logic)
        day_occurrences = {}
        for year in [compare_year, budget_year]:
            for month in range(1, 13):
                for day_num, day_name in enumerate(calendar.day_name):
                    key = (year, month, day_name)
                    day_occurrences[key] = count_day_occurrences(year, month, day_num)
        
        # Group by branch_id, month, day_name - calculate per branch first (matches home page)
        gross_sums = (
            df_compare.groupby(['branch_id', 'month', 'day_name'])['gross']
            .sum()
            .reset_index(name='gross_sum')
        )
        
        # Calculate for each branch-month combination, then aggregate (matches home page logic)
        monthly_data = []
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        for month in range(1, 13):
            month_data = gross_sums[gross_sums['month'] == month]
            
            # Calculate per branch first, then sum
            weekday_aggregated = {}
            
            for day_name in ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
                # Get calendar day counts (same for all branches)
                count_compare = day_occurrences.get((compare_year, month, day_name), 0)
                count_budget = day_occurrences.get((budget_year, month, day_name), 0)
                
                # Calculate per branch, then aggregate
                total_sales_compare = 0
                total_est_sales_budget = 0
                
                # Get data for this day across all branches
                day_data = month_data[month_data['day_name'] == day_name]
                
                for _, row in day_data.iterrows():
                    branch_sales = row['gross_sum']
                    branch_avg = branch_sales / count_compare if count_compare > 0 else 0
                    branch_est = branch_avg * count_budget if count_compare > 0 else 0
                    
                    total_sales_compare += branch_sales
                    total_est_sales_budget += branch_est
                
                # Calculate aggregated average
                avg_compare = total_sales_compare / count_compare if count_compare > 0 else 0
                
                weekday_aggregated[day_name] = {
                    'count_compare': int(count_compare),
                    'sales_compare': float(total_sales_compare),
                    'avg_compare': float(avg_compare),
                    'count_budget': int(count_budget),
                    'sales_budget': 0,  # Not used in calculation
                    'avg_budget': 0,    # Not used in calculation
                    'est_sales_budget': float(total_est_sales_budget)
                }
            
            monthly_data.append({
                'month': month,
                'monthName': month_names[month - 1],
                'weekdayData': weekday_aggregated
            })
        
        return {
            'data': monthly_data,
            'config': {
                'compare_year': compare_year,
                'budget_year': budget_year
            }
        }
        
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to calculate weekend effect: {str(e)}"
        )


@budgetRouter.post("/islamic-calendar-effects", status_code=status.HTTP_200_OK)
def get_islamic_calendar_effects(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Calculate Islamic calendar effects independently (no dependency on home page cache)
    Directly accesses BaseData.pkl and calculates Ramadan, Muharram, and Eid al-Adha impacts
    
    Request body: {
        "branch_ids": [1, 2, 3],
        "compare_year": 2025,
        "budget_year": 2026,
        "ramadan_CY": "2025-03-01",
        "ramadan_BY": "2026-02-18",
        "ramadan_daycount_CY": 30,
        "ramadan_daycount_BY": 30,
        "muharram_CY": "2025-07-27",
        "muharram_BY": "2026-07-16",
        "muharram_daycount_CY": 10,
        "muharram_daycount_BY": 10,
        "eid2_CY": "2025-06-06",
        "eid2_BY": "2026-05-27"
    }
    """
    try:
        import pandas as pd
        import os
        # Import calculation functions from budget.py
        from src.services.budget import (
            Ramadan_Eid_Calculations,
            Muharram_calculations,
            Eid2Calculations,
            Eid2Calculations_v2,
            descriptiveCalculations
        )
        
        # Extract parameters
        branch_ids = request.get("branch_ids", [])
        compare_year = request.get("compare_year", 2025)
        budget_year = request.get("budget_year", 2026)
        
        if not branch_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="branch_ids is required"
            )
        
        # Islamic calendar dates
        ramadan_CY = pd.to_datetime(request.get("ramadan_CY", "2025-03-01"))
        ramadan_BY = pd.to_datetime(request.get("ramadan_BY", "2026-02-18"))
        ramadan_daycount_CY = request.get("ramadan_daycount_CY", 30)
        ramadan_daycount_BY = request.get("ramadan_daycount_BY", 30)
        
        muharram_CY = pd.to_datetime(request.get("muharram_CY", "2025-07-27"))
        muharram_BY = pd.to_datetime(request.get("muharram_BY", "2026-07-16"))
        muharram_daycount_CY = request.get("muharram_daycount_CY", 10)
        muharram_daycount_BY = request.get("muharram_daycount_BY", 10)
        
        eid2_CY = pd.to_datetime(request.get("eid2_CY", "2025-06-06"))
        eid2_BY = pd.to_datetime(request.get("eid2_BY", "2026-05-27"))
        
        # Load BaseData.pkl (using absolute path for reliability)
        # File location: /home/user/backend/Backend/src/api/routes/budget.py
        # Need to go up 4 levels: routes -> api -> src -> Backend
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        pkl_path = os.path.join(base_dir, "BaseData.pkl")
        
        if not os.path.exists(pkl_path):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"BaseData.pkl not found at {pkl_path}"
            )
        
        df = pd.read_pickle(pkl_path)
        
        # DEBUG: Log total data before filtering
        print(f"📊 Total rows in BaseData.pkl: {len(df)}")
        print(f"📊 Date range: {df['business_date'].min()} to {df['business_date'].max()}")
        print(f"📊 Requested branch_ids: {branch_ids}")
        
        # Filter by selected branches
        df = df[df["branch_id"].isin(branch_ids)]
        
        # DEBUG: Log filtered data
        print(f"📊 Rows after branch filter: {len(df)}")
        print(f"📊 Unique branches in filtered data: {df['branch_id'].unique().tolist()}")
        if not df.empty:
            april_2025 = df[(df['business_date'].dt.year == 2025) & (df['business_date'].dt.month == 4)]
            print(f"📊 April 2025 rows: {len(april_2025)}, Total gross: {april_2025['gross'].sum():.2f}")
        
        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No data found for selected branches"
            )
        
        # Calculate Islamic calendar effects
        ramadan = Ramadan_Eid_Calculations(
            compare_year,
            ramadan_daycount_CY, ramadan_daycount_BY,
            ramadan_CY, ramadan_BY,
            df
        )
        
        muh_result = Muharram_calculations(
            compare_year,
            muharram_CY, muharram_BY,
            muharram_daycount_CY, muharram_daycount_BY,
            df
        )
        
        # Extract monthly summary and metadata
        muh = muh_result['monthly_summary']
        muh_metadata = muh_result['metadata']
        
        # Use corrected Eid2Calculations_v2 function
        eid2 = Eid2Calculations_v2(compare_year, eid2_CY, eid2_BY, df)
        
        # Get descriptive statistics (actual sales)
        summarydf = descriptiveCalculations(compare_year, df)
        
        # Merge calculations
        # CRITICAL: Start with summarydf (all 12 months) to ensure all months are included
        # Previously started with ramadan which only had Ramadan-affected months
        final_df = summarydf[['branch_id', 'month', 'total_sales']].copy()
        final_df = pd.merge(final_df, ramadan[['branch_id', 'month', 'Ramadan Eid %', 'est']],
                            on=['branch_id', 'month'], how='left')
        final_df = pd.merge(final_df, muh[['branch_id', 'month', 'Muharram %']],
                            on=['branch_id', 'month'], how='left')
        final_df = pd.merge(final_df, eid2[['branch_id', 'month', 'Eid2 %']],
                            on=['branch_id', 'month'], how='left')
        
        # DEBUG: Check Eid2 data after merge
        print("\n🔍 DEBUG: Eid2 data after merge:")
        print(f"Eid2 DataFrame shape: {eid2.shape}")
        print(f"Eid2 DataFrame dtypes: {eid2.dtypes.to_dict()}")
        print(f"Eid2 data:\n{eid2}")
        print(f"\nFinal_df before merge - branch_ids: {final_df['branch_id'].unique()}, dtype: {final_df['branch_id'].dtype}")
        print(f"Final_df before merge - months: {sorted(final_df['month'].unique())}, dtype: {final_df['month'].dtype}")
        print(f"Eid2 - branch_ids: {eid2['branch_id'].unique()}, dtype: {eid2['branch_id'].dtype}")
        print(f"Eid2 - months: {sorted(eid2['month'].unique())}, dtype: {eid2['month'].dtype}")
        
        # Check if months 5 and 6 exist in final_df
        month5 = final_df[(final_df['branch_id'] == 189) & (final_df['month'] == 5)]
        month6 = final_df[(final_df['branch_id'] == 189) & (final_df['month'] == 6)]
        print(f"\nFinal_df has month 5 for branch 189: {len(month5) > 0}")
        print(f"Final_df has month 6 for branch 189: {len(month6) > 0}")
        
        eid2_rows = final_df[final_df['Eid2 %'].notna()]
        print(f"\nFinal_df rows with Eid2 data AFTER merge: {len(eid2_rows)}")
        if len(eid2_rows) > 0:
            print(eid2_rows[['branch_id', 'month', 'Eid2 %']])
        else:
            print("❌ No Eid2 data found in merged dataframe!")
            sample = final_df[(final_df['branch_id'] == 189) & (final_df['month'] == 5)]
            if len(sample) > 0:
                print(f"Sample row for branch 189, month 5: {sample[['branch_id', 'month', 'Eid2 %']].to_dict('records')}")
        
        # Rename columns for clarity
        final_df = final_df.rename(columns={
            'Ramadan Eid %': 'ramadan_eid_pct',
            'Muharram %': 'muharram_pct',
            'Eid2 %': 'eid2_pct'
        })
        
        # Calculate derived fields
        def calculate_baseline(row):
            ts = row['total_sales']
            if pd.isna(ts) or ts is None:
                return {
                    'sales_CY': None,
                    'est_sales_no_ramadan': None,
                    'est_sales_no_muharram': None,
                    'est_sales_no_eid2': None
                }
            
            muharram_pct = row['muharram_pct']
            eid2_pct = row['eid2_pct']
            
            # CRITICAL: Use 'est' from service layer for Ramadan (correct calculation)
            # Only recalculate for Muharram and Eid2
            return {
                'sales_CY': ts,
                'est_sales_no_ramadan': row['est'] if 'est' in row and not pd.isna(row['est']) else ts,
                'est_sales_no_muharram': ts / (1 + muharram_pct / 100.0) if not pd.isna(muharram_pct) else ts,
                'est_sales_no_eid2': ts / (1 + eid2_pct / 100.0) if not pd.isna(eid2_pct) else ts
            }
        
        # Apply calculations
        derived = final_df.apply(calculate_baseline, axis=1, result_type='expand')
        final_df = pd.concat([final_df, derived], axis=1)
        
        # Get branch details from database
        from src.core.db import get_session, close_session
        from src.db.dbtables import Branch, Brand
        
        dbs = get_session()
        try:
            # Build response structure
            brands_dict = {}
            
            for _, row in final_df.iterrows():
                branch_id = int(row['branch_id'])
                month = int(row['month'])
                
                # Get branch info
                branch = dbs.query(Branch).filter(Branch.id == branch_id).first()
                if not branch:
                    continue
                
                brand_id = branch.brand_id
                brand = dbs.query(Brand).filter(Brand.id == brand_id).first()
                if not brand:
                    continue
                
                # Initialize brand structure
                if brand_id not in brands_dict:
                    brands_dict[brand_id] = {
                        'brand_id': brand_id,
                        'brand_name': brand.name,
                        'branches': {}
                    }
                
                # Initialize branch structure
                if branch_id not in brands_dict[brand_id]['branches']:
                    brands_dict[brand_id]['branches'][branch_id] = {
                        'branch_id': branch_id,
                        'branch_name': branch.name,
                        'months': []
                    }
                
                # Get ACTUAL daily sales for this branch and month from BaseData
                daily_sales_data = []
                branch_month_df = df[
                    (df['branch_id'] == branch_id) & 
                    (df['business_date'].dt.year == compare_year) &
                    (df['business_date'].dt.month == month)
                ]
                
                if not branch_month_df.empty:
                    # 🧠 SMART RAMADAN SYSTEM: Dynamic reference period selection
                    # Initialize smart system (one-time per branch processing)
                    if 'smart_system' not in locals():
                        smart_config = {
                            'compare_year': compare_year,
                            'ramadan_CY': ramadan_CY.strftime('%Y-%m-%d'),
                            'ramadan_BY': ramadan_BY.strftime('%Y-%m-%d'),
                            'ramadan_daycount_CY': ramadan_daycount_CY,
                            'ramadan_daycount_BY': ramadan_daycount_BY
                        }
                        smart_system = SmartRamadanSystem(smart_config)
                        estimation_plan = smart_system.generate_estimation_plan()
                        print(f"\n🎯 Smart System Activated - Dynamic month and day detection")
                    
                    # Pre-calculate weekday averages for all unique reference periods in this month
                    # This is more efficient than recalculating for each day
                    weekday_avg_cache = {}  # Cache: (source_period_key) -> weekday_averages_dict
                    eid_values_cache = {}    # Cache: eid_day_number -> actual_value
                    
                    if month in estimation_plan:
                        # Identify unique reference periods needed for this month
                        unique_references = set()
                        for day_num in estimation_plan[month].keys():
                            ref = estimation_plan[month][day_num]
                            if ref['method'] == 'weekday_average':
                                # Create cache key from source info
                                cache_key = (ref['source_day_type'], tuple(ref['source_months']), str(ref.get('source_date_range')))
                                unique_references.add(cache_key)
                        
                        # Calculate weekday averages for each unique reference period
                        for cache_key in unique_references:
                            source_day_type, source_months, date_range_str = cache_key
                            
                            # Get the first day's reference to extract date range if available
                            sample_ref = estimation_plan[month][1]
                            for day_num, ref in estimation_plan[month].items():
                                ref_key = (ref.get('source_day_type'), tuple(ref.get('source_months', [])), str(ref.get('source_date_range')))
                                if ref_key == cache_key:
                                    sample_ref = ref
                                    break
                            
                            # Filter data based on source period type
                            if sample_ref.get('source_date_range'):
                                # Use specific date range (Ramadan period OR April excluding Eid)
                                start_date, end_date = sample_ref['source_date_range']
                                period_df = df[
                                    (df['branch_id'] == branch_id) & 
                                    (df['business_date'] >= start_date) &
                                    (df['business_date'] <= end_date)
                                ].copy()
                            else:
                                # Normal days: use entire month(s)
                                period_df = df[
                                    (df['branch_id'] == branch_id) & 
                                    (df['business_date'].dt.year == compare_year) &
                                    (df['business_date'].dt.month.isin(source_months))
                                ].copy()
                            
                            if not period_df.empty:
                                period_df['day_of_week'] = period_df['business_date'].dt.day_name()
                                daily_totals = period_df.groupby(['business_date', 'day_of_week'])['gross'].sum().reset_index()
                                weekday_avg_df = daily_totals.groupby('day_of_week')['gross'].mean()
                                weekday_avg_cache[cache_key] = weekday_avg_df.to_dict()
                                
                                period_desc = f"{source_day_type} days from months {source_months}"
                                print(f"📊 Calculated weekday averages for {period_desc} - branch {branch_id}")
                        
                        # Pre-fetch Eid day values if needed for this month
                        for day_num, ref in estimation_plan[month].items():
                            if ref['method'] == 'direct_copy' and ref['eid_day_mapping']:
                                eid_mapping = ref['eid_day_mapping']
                                eid_day_num = eid_mapping['eid_day_number']
                                
                                if eid_day_num not in eid_values_cache:
                                    cy_date = eid_mapping['cy_date']
                                    eid_df = df[
                                        (df['branch_id'] == branch_id) & 
                                        (df['business_date'].dt.year == cy_date.year) &
                                        (df['business_date'].dt.month == cy_date.month) &
                                        (df['business_date'].dt.day == cy_date.day)
                                    ]
                                    if not eid_df.empty:
                                        eid_values_cache[eid_day_num] = float(eid_df['gross'].sum())
                                        print(f"📊 Fetched CY Eid Day {eid_day_num} value: {eid_values_cache[eid_day_num]:.2f} - branch {branch_id}")
                    
                    # Group by day to get actual daily totals
                    daily_totals = branch_month_df.groupby(branch_month_df['business_date'].dt.day)['gross'].sum()
                    
                    # 🐑 INITIALIZE EID AL-ADHA (EID2) CACHE for this branch (one-time)
                    if 'eid2_cache' not in locals():
                        eid2_cache = {}
                        cy_eid_start = pd.to_datetime(eid2_CY)
                        cy_eid_dates = pd.date_range(start=cy_eid_start, periods=3, freq='D')
                        
                        for i, cy_eid_date in enumerate(cy_eid_dates):
                            eid_day_num = i + 1
                            eid_df = df[
                                (df['branch_id'] == branch_id) & 
                                (df['business_date'].dt.year == cy_eid_date.year) &
                                (df['business_date'].dt.month == cy_eid_date.month) &
                                (df['business_date'].dt.day == cy_eid_date.day)
                            ]
                            if not eid_df.empty:
                                eid2_cache[eid_day_num] = float(eid_df['gross'].sum())
                                print(f"🐑 Cached CY Eid2 Day {eid_day_num} ({cy_eid_date.date()}): {eid2_cache[eid_day_num]:.2f} BHD - branch {branch_id}")
                    
                    # Precompute BY Eid dates for quick lookup
                    by_eid_start = pd.to_datetime(eid2_BY)
                    by_eid_dates = pd.date_range(start=by_eid_start, periods=3, freq='D')
                    
                    # 🐑 CALCULATE EID2 WEEKDAY AVERAGES for non-Eid days
                    # Check if this month is affected by Eid2 (either CY or BY has Eid in this month)
                    cy_eid_months = sorted(list(set([d.month for d in cy_eid_dates])))
                    by_eid_months = sorted(list(set([d.month for d in by_eid_dates])))
                    is_eid2_affected_month = month in cy_eid_months or month in by_eid_months
                    
                    eid2_weekday_avg = {}
                    if is_eid2_affected_month:
                        # Get CY month data and exclude CY Eid days
                        cy_month_data_for_avg = df[
                            (df['branch_id'] == branch_id) & 
                            (df['business_date'].dt.year == compare_year) &
                            (df['business_date'].dt.month == month)
                        ].copy()
                        
                        # Exclude CY Eid days from this month
                        cy_eid_days_this_month = [d for d in cy_eid_dates if d.month == month]
                        cy_non_eid_data = cy_month_data_for_avg[~cy_month_data_for_avg['business_date'].isin(cy_eid_days_this_month)]
                        
                        if not cy_non_eid_data.empty:
                            # CRITICAL FIX: First aggregate by day to get DAILY totals, then calculate weekday averages
                            # Group by business_date to get daily totals first
                            cy_daily_totals = cy_non_eid_data.groupby('business_date')['gross'].sum().reset_index()
                            cy_daily_totals['day_of_week'] = cy_daily_totals['business_date'].dt.day_name()
                            
                            # Now calculate weekday averages from daily totals
                            eid2_weekday_avg = cy_daily_totals.groupby('day_of_week')['gross'].mean().to_dict()
                            print(f"🐑 Eid2 weekday averages for month {month} calculated from {len(cy_daily_totals)} non-Eid days")
                            print(f"   Sample averages: {list(eid2_weekday_avg.items())[:3]}")
                    
                    # 🟠 CHECK: Is this a Muharram-affected month?
                    is_muharram_month = month in muh_metadata['affected_months_BY']
                    
                    for day_num, daily_gross in daily_totals.items():
                        estimated_value = float(daily_gross)  # Default fallback
                        
                        # Determine BY date for this day
                        date_BY = pd.Timestamp(year=budget_year, month=month, day=day_num)
                        
                        # 🐑 HIGHEST PRIORITY: Eid al-Adha (Eid2) direct copy
                        if date_BY in by_eid_dates:
                            # This is a BY Eid day - copy from CY Eid
                            by_eid_day_num = (date_BY - by_eid_start).days + 1
                            if by_eid_day_num in eid2_cache:
                                estimated_value = float(eid2_cache[by_eid_day_num])
                                print(f"   🐑 BY Eid2 Day {by_eid_day_num} ({date_BY.date()}): Copied {estimated_value:.2f} BHD")
                        
                        # 🐑 SECOND PRIORITY: Use Eid2 weekday averages for non-Eid days in affected months
                        elif is_eid2_affected_month and eid2_weekday_avg:
                            # This is a non-Eid day in an Eid2-affected month
                            day_of_week_BY = date_BY.day_name()
                            if day_of_week_BY in eid2_weekday_avg:
                                estimated_value = float(eid2_weekday_avg[day_of_week_BY])
                                print(f"   🐑 BY Non-Eid Day {day_num} ({date_BY.date()}, {day_of_week_BY}): Using weekday avg {estimated_value:.2f} BHD")
                            else:
                                estimated_value = float(daily_gross)  # Fallback to CY actual
                        
                        # 🟠 PRIORITY: Use Muharram TWO separate weekday averages for affected months
                        elif is_muharram_month and branch_id in muh_metadata['branch_weekday_averages']:
                            # Get Muharram weekday averages for this branch
                            branch_avg = muh_metadata['branch_weekday_averages'][branch_id]
                            
                            # Determine day of week for BY date
                            day_of_week_BY = date_BY.day_name()
                            
                            # Determine if this day falls within Muharram period in BY 2026
                            muharram_start_BY = muh_metadata['muharram_start_BY']
                            muharram_end_BY = muh_metadata['muharram_end_BY']
                            is_muharram_day = (muharram_start_BY <= date_BY <= muharram_end_BY)
                            
                            # Use appropriate weekday average based on whether day is in Muharram period
                            if is_muharram_day:
                                # Use MUHARRAM weekday average
                                estimated_value = float(branch_avg['MUHARRAM'].get(day_of_week_BY, daily_gross))
                            else:
                                # Use NON-MUHARRAM weekday average
                                estimated_value = float(branch_avg['NON_MUHARRAM'].get(day_of_week_BY, daily_gross))
                        
                        # 🧠 FALLBACK: Use Ramadan Smart System for non-Muharram, non-Eid2 months
                        if estimated_value == float(daily_gross) and month in estimation_plan and day_num in estimation_plan[month]:
                            ref = estimation_plan[month][day_num]
                            
                            # Determine day of week for BY date (for weekday averaging)
                            date_BY = pd.Timestamp(year=budget_year, month=month, day=day_num)
                            day_of_week_BY = date_BY.day_name()
                            
                            if ref['method'] == 'direct_copy':
                                # RULE 1: Eid days - direct copy from CY Eid day
                                eid_mapping = ref['eid_day_mapping']
                                eid_day_num = eid_mapping['eid_day_number']
                                if eid_day_num in eid_values_cache:
                                    estimated_value = float(eid_values_cache[eid_day_num])
                                else:
                                    estimated_value = float(daily_gross)  # Fallback to actual
                            
                            elif ref['method'] == 'weekday_average':
                                # RULE 2 & 3: Ramadan or Normal days - weekday averaging
                                cache_key = (ref['source_day_type'], tuple(ref['source_months']), str(ref.get('source_date_range')))
                                if cache_key in weekday_avg_cache:
                                    weekday_averages = weekday_avg_cache[cache_key]
                                    estimated_value = float(weekday_averages.get(day_of_week_BY, daily_gross))
                                else:
                                    estimated_value = float(daily_gross)  # Fallback to actual
                        
                        daily_sales_data.append({
                            'day': int(day_num),
                            'actual': float(daily_gross),
                            'estimated': estimated_value
                        })
                
                # Helper function to safely convert to float, replacing inf/nan with None
                def safe_float(value):
                    if pd.isna(value):
                        return None
                    f = float(value)
                    if f == float('inf') or f == float('-inf'):
                        return 0
                    return f
                
                # Calculate daily_sales total for comparison
                daily_total = sum(day['estimated'] for day in daily_sales_data) if daily_sales_data else 0
                
                # DEBUG: Check for mismatches
                if branch_id == 189 and month == 4:
                    print(f"\n🔍 DEBUG April 2026 - Branch 189:")
                    print(f"   est_sales_no_ramadan (tile): {row['est_sales_no_ramadan']:.2f}")
                    print(f"   daily_sales total (table): {daily_total:.2f}")
                    print(f"   Difference: {abs(row['est_sales_no_ramadan'] - daily_total):.2f}")
                
                # Add month data with daily breakdown
                month_data = {
                    'month': month,
                    'sales_CY': safe_float(row['sales_CY']),
                    'est_sales_no_ramadan': safe_float(row['est_sales_no_ramadan']),
                    'est_sales_no_muharram': safe_float(row['est_sales_no_muharram']),
                    'est_sales_no_eid2': safe_float(row['est_sales_no_eid2']),
                    'ramadan_eid_pct': safe_float(row['ramadan_eid_pct']),
                    'muharram_pct': safe_float(row['muharram_pct']),
                    'eid2_pct': safe_float(row['eid2_pct']),
                    'daily_sales': daily_sales_data  # Add actual daily sales array
                }
                
                brands_dict[brand_id]['branches'][branch_id]['months'].append(month_data)
            
            # 🌙 ADD MUHARRAM MONTHS (July=7, August=8) with daily sales data
            print("\n🌙 Adding Muharram months (July/August) to API response...")
            # Process each branch that was already added above
            for brand_id in brands_dict.keys():
                for branch_id in brands_dict[brand_id]['branches'].keys():
                    # 🌙 MUHARRAM FULL ISLAMIC MONTH ANALYSIS
                    # CY 2025: Muharram Islamic month spans July 27 - August 5 (10 days)
                    # BY 2026: Muharram Islamic month spans July 16 - July 25 (10 days)
                    # Strategy: Calculate TWO weekday averages from CY 2025:
                    #   1. Muharram days weekday average (from the 10 Islamic days)
                    #   2. Non-Muharram days weekday average (from remaining days in July+August)
                    
                    # Define Muharram dates (from API request parameters)
                    muharram_start_CY = muharram_CY  # 2025-07-27
                    muharram_end_CY = muharram_CY + timedelta(days=muharram_daycount_CY - 1)  # 2025-08-05
                    muharram_start_BY = muharram_BY  # 2026-07-16
                    muharram_end_BY = muharram_BY + timedelta(days=muharram_daycount_BY - 1)  # 2026-07-25
                    
                    print(f"🌙 Processing Muharram Islamic month for branch {branch_id}...")
                    print(f"   CY 2025 Muharram: {muharram_start_CY.strftime('%b %d')} - {muharram_end_CY.strftime('%b %d')}")
                    print(f"   BY 2026 Muharram: {muharram_start_BY.strftime('%b %d')} - {muharram_end_BY.strftime('%b %d')}")
                    
                    # Get ALL data for June + July 2025 (CY) - the Muharram-affected months
                    branch_june_july_df = df[
                        (df['branch_id'] == branch_id) & 
                        (df['business_date'].dt.year == compare_year) &
                        (df['business_date'].dt.month.isin([6, 7]))
                    ].copy()
                    
                    if not branch_june_july_df.empty:
                        # Calculate TWO separate weekday averages (NON-MUHARRAM and MUHARRAM)
                        print(f"   📊 CY 2025: Processing FULL months - {len(branch_june_july_df)} total days")
                        
                        # Separate NON-MUHARRAM and MUHARRAM days
                        branch_june_july_df['day_of_week'] = branch_june_july_df['business_date'].dt.day_name()
                        
                        # NON-MUHARRAM days (June 1-25 + July 26-31 in CY 2025)
                        non_muharram_df = branch_june_july_df[
                            ~branch_june_july_df['business_date'].between(muharram_start_CY, muharram_end_CY)
                        ].copy()
                        
                        # MUHARRAM days (June 26 - July 25 in CY 2025)
                        muharram_df = branch_june_july_df[
                            branch_june_july_df['business_date'].between(muharram_start_CY, muharram_end_CY)
                        ].copy()
                        
                        print(f"   🟢 Non-Muharram days: {len(non_muharram_df)}")
                        print(f"   🟠 Muharram days: {len(muharram_df)}")
                        
                        # Calculate NON-MUHARRAM weekday averages
                        weekday_avg_NON_MUHARRAM = {}
                        if not non_muharram_df.empty:
                            daily_totals_non = non_muharram_df.groupby(['business_date', 'day_of_week'])['gross'].sum().reset_index()
                            weekday_avg_NON_MUHARRAM = daily_totals_non.groupby('day_of_week')['gross'].mean().to_dict()
                            print(f"   🌙 NON-MUHARRAM weekday averages: {weekday_avg_NON_MUHARRAM}")
                        
                        # Calculate MUHARRAM weekday averages
                        weekday_avg_MUHARRAM = {}
                        if not muharram_df.empty:
                            daily_totals_muh = muharram_df.groupby(['business_date', 'day_of_week'])['gross'].sum().reset_index()
                            weekday_avg_MUHARRAM = daily_totals_muh.groupby('day_of_week')['gross'].mean().to_dict()
                            print(f"   🌙 MUHARRAM weekday averages: {weekday_avg_MUHARRAM}")
                    
                    # Now process each month (June and July) for display
                    for muharram_month in [6, 7]:
                        print(f"\n📅 Building daily breakdown for month {muharram_month}...")
                        daily_sales_data = []
                        
                        branch_month_df = df[
                            (df['branch_id'] == branch_id) & 
                            (df['business_date'].dt.year == compare_year) &
                            (df['business_date'].dt.month == muharram_month)
                        ]
                        
                        if not branch_month_df.empty:
                            daily_totals_cy = branch_month_df.groupby(branch_month_df['business_date'].dt.day)['gross'].sum()
                            
                            for day_num, daily_gross_cy in daily_totals_cy.items():
                                date_obj_cy = pd.Timestamp(year=compare_year, month=muharram_month, day=day_num)
                                date_obj_by = pd.Timestamp(year=compare_year + 1, month=muharram_month, day=day_num)
                                day_of_week_by = date_obj_by.day_name()
                                
                                # Determine if BY 2026 day falls within Muharram period
                                is_muharram_day_by = (muharram_start_BY <= date_obj_by <= muharram_end_BY)
                                
                                # Use appropriate weekday average based on Muharram period
                                if is_muharram_day_by:
                                    # Use MUHARRAM weekday average
                                    estimated_value = weekday_avg_MUHARRAM.get(day_of_week_by, float(daily_gross_cy))
                                else:
                                    # Use NON-MUHARRAM weekday average
                                    estimated_value = weekday_avg_NON_MUHARRAM.get(day_of_week_by, float(daily_gross_cy))
                                
                                daily_sales_data.append({
                                    'day': int(day_num),
                                    'actual': float(daily_gross_cy),
                                    'estimated': float(estimated_value)
                                })
                        
                        # Get muharram_pct from service layer
                        muharram_pct = None
                        muharram_row = muh[
                            (muh['branch_id'] == branch_id) & 
                            (muh['month'] == muharram_month)
                        ]
                        if not muharram_row.empty:
                            muharram_pct = safe_float(muharram_row.iloc[0]['Muharram %'])
                        
                        # Calculate totals
                        sales_CY = sum(day['actual'] for day in daily_sales_data) if daily_sales_data else 0
                        est_sales_BY = sum(day['estimated'] for day in daily_sales_data) if daily_sales_data else 0
                        
                        # Recalculate percentage based on full month
                        if sales_CY > 0:
                            muharram_pct = ((est_sales_BY - sales_CY) / sales_CY) * 100
                        
                        est_sales_no_muharram = sales_CY / (1 + muharram_pct / 100.0) if muharram_pct != 0 else sales_CY
                        
                        month_data = {
                            'month': muharram_month,
                            'sales_CY': sales_CY,
                            'est_sales_no_ramadan': sales_CY,
                            'est_sales_no_muharram': est_sales_no_muharram,
                            'est_sales_no_eid2': sales_CY,
                            'ramadan_eid_pct': 0.0,
                            'muharram_pct': muharram_pct if muharram_pct is not None else 0.0,
                            'eid2_pct': 0.0,
                            'daily_sales': daily_sales_data
                        }
                        
                        brands_dict[brand_id]['branches'][branch_id]['months'].append(month_data)
                        print(f"✅ Month {muharram_month}: {len(daily_sales_data)} days, CY: {sales_CY:.2f}, BY est: {est_sales_BY:.2f}, Impact: {muharram_pct:.2f}%")
            
            # Convert to list structure
            result = []
            for brand_data in brands_dict.values():
                brand_data['branches'] = list(brand_data['branches'].values())
                result.append(brand_data)
            
            return {'data': result}
            
        finally:
            close_session(dbs)
        
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to calculate Islamic calendar effects: {str(e)}"
        )


@budgetRouter.post("/home", status_code=status.HTTP_202_ACCEPTED)
def defaultCalculationHome(current_user: dict = Depends(get_current_user)):
    try:
        # res=calculateDefault(request)
        # return {"Data":res}
        result = compute_or_reuse(None, recompute=calculateDefault)
        # Filter results based on user permissions
        return filter_budget_data_by_permissions(result, current_user)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@budgetRouter.post("/projection-inputs/upsert")
def upsert_projection(body: ProjectionInputModel):
    upsert_projection_input(body.model_dump())
    return {"status": "ok"}


@budgetRouter.post("/projection-inputs/upsert-batch")
def upsert_projection_batch(body: list[dict]):
    """
    Batch upsert multiple projection inputs in a single transaction.
    Body: [{"branch_id": 231, "month": 2, ...}, {"branch_id": 231, "month": 3, ...}]
    """
    try:
        upsert_projection_input_batch(body)
        return {"status": "ok", "count": len(body)}
    except Exception as e:
        print(e)
        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Failed to batch upsert projection inputs: {str(e)}")


@budgetRouter.post("/pre-estimation/upsert")
def upsert_estimate(body: ProjectionEstimateIn):
    try:
        res = upsert_projection_estimate(body.model_dump())
        return {"ok": True, **res}
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(e)
        raise HTTPException(
            status_code=500, detail="Failed to upsert projection estimate")


@budgetRouter.post("/pre-estimation/upsert-batch")
def upsert_estimate_batch(body: list[dict]):
    """
    Batch upsert multiple projection estimates in a single transaction.
    Body: [{"branch_id": 231, "month": 2, ...}, {"branch_id": 231, "month": 3, ...}]
    """
    try:
        upsert_projection_estimate_batch(body)
        return {"status": "ok", "count": len(body)}
    except Exception as e:
        print(e)
        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Failed to batch upsert projection estimates: {str(e)}")


@budgetRouter.post("/allocate-grand-total", response_model=AllocateGrandTotalOut)
def allocate_total_sales(payload: AllocateGrandTotalIn, include_data: bool = True, current_user: dict = Depends(get_current_user)):
    """
    POST /allocation/grand-total?include_data=true
    Body: { "grand_total_estimated_sales": 10000000 }
    Returns:
      - summary (factor, counts…)
      - and, when include_data=true, "data": [ { brand_id, brand_name, branches: [...] } ]
        Each months[] object includes your existing fields PLUS est_* from projection_estimate_adjusted.
    """
    try:
        result = allocate_grand_total(payload, include_data=True)
        return filter_budget_data_by_permissions(result, current_user)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Allocation failed")


@budgetRouter.post("/allocation-monthly-totals", response_model=MonthlyTotalsOut)
def allocate_monthly(payload: MonthlyTotalsIn, include_data: bool = True, current_user: dict = Depends(get_current_user)):
    """
    POST /allocation-monthly-totals?include_data=true
    Body:
    {
      "month_totals": {
        "1": 1000000,
        "2": 900000,
        "3": 950000
      }
    }
    - Only the provided months are re-allocated.
    - Each month uses its own factor based on that month’s baseline sum.
    - Results upsert into projection_estimate_adjusted.
    - When include_data=true, returns the nested data array.
    """
    try:
        result = allocate_monthly_totals(payload, include_data=include_data)
        return filter_budget_data_by_permissions(result, current_user)
    except Exception:
        raise HTTPException(
            status_code=500, detail="Monthly allocation failed")


@budgetRouter.post("/branch-totals", response_model=BranchTotalsOut)
def allocate_branch_totals_api(payload: BranchTotalsIn, include_data: bool = True, current_user: dict = Depends(get_current_user)):
    """
    POST /allocation/branch-totals?include_data=true
    Body: {"branch_totals": {"189": 1200000, "190": null}}
      - 189 scaled to 1.2M across all months
      - 190 copied from baseline (factor=1)
      - Any branch not mentioned is also copied from baseline (factor=1)
    """
    try:
        result = allocate_branch_totals(payload, include_data=include_data)
        return filter_budget_data_by_permissions(result, current_user)
    except Exception:
        raise HTTPException(
            status_code=500, detail="Branch grand-total allocation failed")


@budgetRouter.post("/branch-monthly-totals", response_model=BranchMonthlyTotalsOut)
def allocate_branch_monthly_totals_api(payload: BranchMonthlyTotalsIn, include_data: bool = True, current_user: dict = Depends(get_current_user)):
    """
    POST /allocation/branch-monthly-totals?include_data=true
    Body:
    {
      "branch_month_totals": {
        "189": {"1": 100000, "2": 120000, "3": null},
        "190": {"1": 90000}
      }
    }
    - Provided (branch, month) with value -> scaled to that total
    - Provided (branch, month) with null/<=0 -> copied from baseline (factor=1)
    - Any (branch, month) not provided -> copied from baseline (factor=1)
    """
    try:
        result = allocate_branch_monthly_totals(payload, include_data=include_data)
        return filter_budget_data_by_permissions(result, current_user)
    except Exception:
        raise HTTPException(
            status_code=500, detail="Branch monthly allocation failed")
    
@budgetRouter.get("/branches", response_model=BranchListOut)
def get_branches(brand_id: Optional[int] = Query(None, description="Filter by brand id")):
    """
    GET /branches
    GET /branches?brand_id=38

    Returns:
      {
        "branches": [
          { "branch_id": 189, "branch_name": "AL ABRAAJ SEHLA", "brand_id": 38, "brand_name": "AL ABRAAJ RESTAURANTS" },
          ...
        ]
      }
    """
    try:
        data = list_branches(brand_id=brand_id)
        return {"branches": data}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to list branches")

@budgetRouter.post("/import-sales")
async def import_sales(file: UploadFile = File(...)):
    """
    Validates and imports uploaded sales data (CSV or Excel format) for:
    - exact header (order + case)
    - non-empty, unique OrderID
    - non-empty OrderType
    - non-empty branch_id
    
    Supported formats: .csv, .xlsx, .xls
    Returns human-readable errors or 'ok' if valid.
    """
    ok, errors_or_msg = await validate_sales_csv(file)

    if not ok:
        return JSONResponse(
            status_code=400,
            content={
                "status": "error",
                "message": "File validation failed. Please fix the issues below and try again.",
                "errors": errors_or_msg,
            },
        )

    return {"status": "ok", "message": "Sales data imported successfully."}

@budgetRouter.get("/download-sales-template")
async def download_sales_template():
    """
    Download master Excel template for sales data import.
    
    Returns an Excel file with:
    - Correct column headers in exact order
    - Sample data rows to demonstrate format
    - Professional formatting with colors and borders
    """
    # Define the expected columns
    columns = [
        "OrderID",
        "OrderDateTime",
        "OrderType",
        "branch_id",
        "SubTotal",
        "VAT",
        "OrderDiscount",
        "AmountDue",
        "guests",
        "ItemDiscount",
    ]
    
    # Create sample data (3 example rows)
    sample_data = [
        {
            "OrderID": "ORD001",
            "OrderDateTime": "01/01/2025 14:30",
            "OrderType": "Dinein",
            "branch_id": "189",
            "SubTotal": 45.50,
            "VAT": 2.28,
            "OrderDiscount": 0.00,
            "AmountDue": 47.78,
            "guests": 3,
            "ItemDiscount": 0.00
        },
        {
            "OrderID": "ORD002",
            "OrderDateTime": "01/01/2025 15:45",
            "OrderType": "1",
            "branch_id": "190",
            "SubTotal": 32.00,
            "VAT": 1.60,
            "OrderDiscount": 2.00,
            "AmountDue": 31.60,
            "guests": 0,
            "ItemDiscount": 0.50
        },
        {
            "OrderID": "ORD003",
            "OrderDateTime": "02/01/2025 12:15",
            "OrderType": "2",
            "branch_id": "189",
            "SubTotal": 28.75,
            "VAT": 1.44,
            "OrderDiscount": 0.00,
            "AmountDue": 30.19,
            "guests": 0,
            "ItemDiscount": 0.00
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sales Data')
        
        # Get workbook and worksheet for styling
        workbook = writer.book
        worksheet = writer.sheets['Sales Data']
        
        # Apply professional styling
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Data styling
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Style header row
        for col_num, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            # Set column width
            worksheet.column_dimensions[cell.column_letter].width = 15
        
        # Style data rows with alternating colors
        for row_num in range(2, len(sample_data) + 2):
            fill_color = "F2F2F2" if row_num % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            for col_num in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = data_alignment
        
        # Add instructions sheet
        instructions_data = {
            "Field": columns,
            "Description": [
                "Unique order identifier (required, must be unique)",
                "Date and time of order (format: DD/MM/YYYY HH:MM)",
                "Order type: Dinein, 1=Delivery, 2=Takeaway, 4=Drive Thru, 6=Catering, 7=Staff Meal",
                "Branch ID number (required)",
                "Subtotal amount before tax and discounts",
                "VAT/Tax amount",
                "Order-level discount amount",
                "Final amount due after all calculations",
                "Number of guests (0 for non-dine-in orders)",
                "Item-level discount amount"
            ],
            "Required": [
                "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "No", "No"
            ]
        }
        instructions_df = pd.DataFrame(instructions_data)
        instructions_df.to_excel(writer, index=False, sheet_name='Instructions')
        
        # Style instructions sheet
        inst_worksheet = writer.sheets['Instructions']
        for col_num in range(1, 4):
            cell = inst_worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths for instructions
        inst_worksheet.column_dimensions['A'].width = 20
        inst_worksheet.column_dimensions['B'].width = 70
        inst_worksheet.column_dimensions['C'].width = 12
        
        # Style instructions data rows
        for row_num in range(2, len(columns) + 2):
            for col_num in range(1, 4):
                cell = inst_worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    output.seek(0)
    
    # Return as downloadable file
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Sales_Import_Template.xlsx"
        }
    )